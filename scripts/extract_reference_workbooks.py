from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


SOURCE_DIR = Path("/home/ubuntu/upload")
OUTPUT_DIR = Path("/home/ubuntu/rooted-fit/docs/reference_extraction")
WORKBOOKS = (
    "7-DayPostpartumFlat-StomachMealPlanV2",
    "FoodTimetable",
)


def extract_workbook(source: Path, destination: Path) -> None:
    workbook = load_workbook(BytesIO(source.read_bytes()), read_only=True, data_only=True)
    lines: list[str] = []
    for worksheet in workbook.worksheets:
        lines.append(f"# {worksheet.title}")
        for row in worksheet.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value is not None and str(value).strip()]
            if values:
                lines.append(" | ".join(values))
    destination.write_text("\n".join(lines), encoding="utf-8")


OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
for workbook_name in WORKBOOKS:
    extract_workbook(SOURCE_DIR / workbook_name, OUTPUT_DIR / f"{workbook_name}.md")
